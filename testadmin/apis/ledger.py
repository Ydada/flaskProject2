import io, json
from datetime import datetime

from flask import Blueprint, request, make_response
from flask_jwt_extended import current_user, jwt_required
from flask_sqlalchemy.pagination import Pagination
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from sqlalchemy import inspect

from testadmin.extensions import db
from testadmin.orms import LedgerORM

ledger_api = Blueprint("ledger", __name__)


@ledger_api.get("/ledger")
def ledger_list():
    page = request.args.get("page", default=1, type=int)
    per_page = request.args.get("limit", default=10, type=int)
    q = db.select(LedgerORM)

    pages: Pagination = db.paginate(q, page=page, per_page=per_page)
    print(pages)
    return {
        "code": 0,
        "msg": "获取虚拟机数据成功",
        "data": [item.json() for item in pages.items],
        "count": pages.total,
    }


@ledger_api.post("/ledger")
def ledger_info():
    data = request.get_json()
    if data["id"]:
        del data["id"]
    role = LedgerORM(**data)
    create_at = data["create_at"]
    if create_at:
        role.create_at = datetime.strptime(create_at, "%Y-%m-%d %H:%M:%S")
    role.save()
    return {"code": 0, "msg": "新增虚拟机成功"}


@ledger_api.put("/ledger/<int:uid>")
def change_ledger(uid):
    data = request.get_json()
    del data["id"]

    ledger_obj = LedgerORM.query.get(uid)
    for key, value in data.items():
        if key == "create_at":
            value = datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
        setattr(ledger_obj, key, value)
    ledger_obj.save()
    return {"code": 0, "msg": "修改虚拟机信息成功"}


@ledger_api.delete("/ledger/<int:rid>")
def del_ledger(rid):
    ledger_obj = LedgerORM.query.get(rid)
    ledger_obj.delete()
    return {"code": 0, "msg": "删除虚拟机成功"}


@ledger_api.get("/exportfile")
def export_file():
    qs = LedgerORM.query.all()
    # 创建一个新的工作簿和工作表
    data_tuples = [(q.id, q.hostname, q.ip1, q.ip2, q.ip3, q.ip4, q.app_name, q.app_vendors, q.app_owner, q.app_dest,
                    q.remark, q.create_at) for q in qs]
    print(data_tuples)
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # 写入标题行（可选）
    ws.append(
        ["id", "主机名", "综合网IP地址", "数据网IP地址", "存储网IP地址", "其他IP地址", "应用服务名", "厂商", "负责人",
         "业务内容", "备注", "创建时间"])  # 根据你的数据库模型调整列名

    # 遍历数据并写入工作表
    for item in data_tuples:
        print(item)
        row_data = [item[0], item[1], item[2], item[3], item[4], item[5], item[6], item[7], item[8], item[9], item[10],
                    item[11]]  # 假设你的模型有这些属性
        ws.append(row_data)

        # 将工作簿保存为字节流
    output = io.BytesIO()
    wb.save(output)

    # 将输出指针移回字节流的开头
    output.seek(0)

    # 设置响应头并发送文件
    response = make_response(output.read())
    response.headers["Content-Disposition"] = "attachment; filename=domain.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return response


def get_column_names_inspect(model):
    mapper = inspect(model)
    return [c.key for c in mapper.columns]


@ledger_api.post("/importfile")
def import_file():
    # 数据组装
    datas = []
    # 获取文件对象
    file = request.files['file']
    # 获取文件名
    filename = file.filename
    # file.save 也可保存传来的文件
    # file.save(f'./{filename}')
    with open(f'./static/file/{filename}', 'wb') as f:
        f.write(file.stream.read())
        f.close()
    if filename.endswith('.xlsx'):
        wb = load_workbook(filename=f'./static/file/{filename}')
        sheet_names = wb.sheetnames
        sheet = wb[sheet_names[0]]
        rows = sheet.max_row  # 获取行数
        cols = sheet.max_column  # 获取列数
        # data = request.json  # 假设前端发送的是 JSON 数据
        # 获取表头
        head = get_column_names_inspect(LedgerORM)
        print(head)
        for row in sheet.iter_rows(min_row=2, max_row=rows + 1, values_only=True):
            data = dict(zip(head, row))
            datas.append(data)
        # # 将数据列表转换为 JSON 字符串
        # json_data = json.dumps(data, indent=4, ensure_ascii=False)
        # 将 JSON 字符串写入文件
        # with open('output.json', 'w', encoding='utf-8') as f:
        #     f.write(datas)
        # 或者打印到控制台
        print(datas)
        # for i in range(2, rows):
        #     for j in range(2, cols + 1):
        #         cell = sheet.cell(i, j).value
        #         print(cell, end=" ")
        #         if cell is not None:
        #             existing_record = LedgerORM.query.filter_by(unique_field=data['unique_field']).first()
        #             if existing_record:
        #                 # 更新记录
        #                 existing_record.some_field = data['some_field']  # 假设你要更新 some_field
        #                 db.session.commit()
        #                 return "Record updated successfully."
        #             else:
        #                 # 插入新记录
        #                 new_record = LedgerORM(unique_field=data['unique_field'], some_field=data['some_field'])
        #                 db.session.add(new_record)
        #                 db.session.commit()
        #                 return "New record inserted successfully."
        #
        #     print()
    return {
        "code": 0,
        "msg": "获取虚拟机数据成功",
        "data": datas,
        "count": len(datas),
    }


@ledger_api.get("/getimportexcel")
def get_import_excel():
    # 数据组装
    datas = []
    # 获取文件对象
    file = request.files['file']
    # 获取文件名
    filename = file.filename
    # file.save 也可保存传来的文件
    # file.save(f'./{filename}')
    with open(f'./static/file/{filename}', 'wb') as f:
        f.write(file.stream.read())
        f.close()
    if filename.endswith('.xlsx'):
        wb = load_workbook(filename=f'./static/file/{filename}')
        sheet_names = wb.sheetnames
        sheet = wb[sheet_names[0]]
        rows = sheet.max_row  # 获取行数
        cols = sheet.max_column  # 获取列数
        # 获取表头
        head = get_column_names_inspect(LedgerORM)
        for row in sheet.iter_rows(min_row=2, max_row=rows + 1, values_only=True):
            data = dict(zip(head, row))
            datas.append(data)
    return {
        "code": 0,
        "msg": "获取虚拟机数据成功",
        "data": datas,
        "count": len(datas),
    }


#
# @user_api.get("/user/user_role/<int:uid>")
# def get_user_role(uid):
#     user: UserORM = db.session.execute(
#         db.select(UserORM).where(UserORM.id == uid)
#     ).scalar()
#
#     wn_role_list = [r.id for r in user.role_list]
#
#     return {
#         "code": 0,
#         "msg": "返回角色权限数据成功",
#         "data": wn_role_list,
#     }
#
#
# @user_api.put("/user/user_role/<int:rid>")
# def change_user_role(rid):
#     role_ids = request.json.get("rights_ids", "")
#     role_list = role_ids.split(",")
#
#     user: UserORM = db.session.execute(
#         db.select(UserORM).where(UserORM.id == rid)
#     ).scalar()
#     role_obj_list = db.session.execute(
#         db.select(RoleORM).where(RoleORM.id.in_(role_list))
#     ).all()
#     user.role_list = [r[0] for r in role_obj_list]
#     user.save()
#     return {"code": 0, "msg": "授权成功"}


@ledger_api.get("/user/profile")
@jwt_required()
def user_profile():
    return {
        "code": 0,
        "msg": "获取个人数据成功",
        "data": current_user.json(),
    }
